import os, sys, subprocess, csv as csv_module, re, json
from pathlib import Path
from datetime import datetime
from flask import Flask, request, jsonify, render_template_string, send_from_directory

app = Flask(__name__, static_folder=None, static_url_path=None)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024


@app.after_request
def add_cors(resp):
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return resp

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_RAW = BASE_DIR / "data" / "raw"
SCRIPTS_DIR = BASE_DIR / "scripts"

DATA_RAW.mkdir(parents=True, exist_ok=True)

ALLOWED_EXT = {".csv", ".xlsx", ".xls"}


def _is_allowed(filename):
    return Path(filename).suffix.lower() in ALLOWED_EXT


def _xlsx_to_csv(xlsx_path, csv_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    data_start = 0
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c is not None and str(c).strip() != ""]
        if len(non_empty) <= 1:
            data_start = i + 1
        else:
            break
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv_module.writer(f, delimiter=";")
        for row in rows[data_start:]:
            w.writerow(["" if v is None else v for v in row])
    return csv_path


def _run_pipeline(csv_path):
    stem = Path(csv_path).stem
    date_stamp = datetime.now().strftime("%Y-%m-%d")
    report_dir = BASE_DIR / "reports" / stem
    report_dir.mkdir(parents=True, exist_ok=True)

    proc = subprocess.run(
        [sys.executable, str(SCRIPTS_DIR / "run_pipeline.py"), str(csv_path), "-o", str(report_dir)],
        capture_output=True, text=True, cwd=str(BASE_DIR),
    )

    stdout = proc.stdout or ""
    stderr = proc.stderr or ""

    if proc.returncode != 0:
        raise RuntimeError(f"Pipeline failed:\n{stderr[:1000]}")

    reports = {}
    for line in stdout.split("\n"):
        m = re.match(r"\s*-\s*(Summary|Dashboard|PDF):\s+(.+)", line)
        if m:
            key = m.group(1).lower()
            raw_path = m.group(2).strip()
            raw_path = raw_path.replace("\\\\", "/").replace("\\", "/")
            parts = raw_path.split("reports/", 1)
            rel = "reports/" + parts[1] if len(parts) > 1 else raw_path
            reports[key] = rel

    return {
        "summary": reports.get("summary", ""),
        "dashboard": reports.get("dashboard", ""),
        "pdf": reports.get("pdf", ""),
    }


@app.route("/")
def index():
    return send_from_directory(str(Path(__file__).parent), "index.html")


@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify(status="error", message="Файл не выбран"), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify(status="error", message="Файл не выбран"), 400
    if not _is_allowed(file.filename):
        return jsonify(status="error", message="Допустимы только файлы CSV, XLSX, XLS"), 400

    ext = Path(file.filename).suffix.lower()
    raw_path = DATA_RAW / file.filename
    file.save(str(raw_path))

    try:
        if ext == ".csv":
            csv_path = raw_path
        else:
            csv_name = Path(file.filename).stem + ".csv"
            csv_path = DATA_RAW / csv_name
            _xlsx_to_csv(str(raw_path), str(csv_path))

        reports = _run_pipeline(str(csv_path))
        return jsonify(
            status="ok",
            file_name=Path(csv_path).name,
            message="Файл успешно загружен и обработан",
            reports=reports,
        )
    except RuntimeError as e:
        return jsonify(status="error", message=str(e)), 500
    except Exception as e:
        return jsonify(status="error", message=f"Ошибка: {str(e)}"), 500


@app.route("/reports/<path:filename>")
def serve_report(filename):
    return send_from_directory(str(BASE_DIR / "reports"), filename)


@app.errorhandler(413)
def too_large(e):
    return jsonify(status="error", message="Файл превышает максимальный размер 50 MB"), 413


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)
